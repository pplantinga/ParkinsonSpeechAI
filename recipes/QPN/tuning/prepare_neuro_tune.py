import json
import os
import glob
import torchaudio
import openpyxl
import numpy as np
import pathlib


def convert_to_python(obj):
    """Recursively convert NumPy types to Python types for JSON serialization."""
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {k: convert_to_python(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_python(item) for item in obj]
    else:
        return obj


def prepare_neuro(
    data_folder, train_annotation, valid_annotation, chunk_size, split):
    assert os.path.exists(data_folder), "Data folder not found"

    # Gather ALL recordings into one dict, ignoring train/valid folder structure
    all_path_type_dict = get_all_path_type_dict(data_folder)

    # Deterministic 50/50 stratified split by patient ID and ptype
    split_a, split_b = stratified_patient_split(all_path_type_dict)

    # The split variable (0 or 1) determines which half is train vs. valid
    if split == 0:
        train_dict, valid_dict = split_a, split_b
    else:
        train_dict, valid_dict = split_b, split_a

    create_json(train_annotation, train_dict, chunk_size, overlap=None)
    create_json(valid_annotation, valid_dict, chunk_size)


def get_all_path_type_dict(data_folder):
    """
    Collects ALL recordings from the data folder into a single dict,
    irrespective of train/valid subdirectory structure.

    :param data_folder: string
    :return: dict of {path: patient_traits}
    """
    batch1_excel_path = os.path.join(data_folder, "QPN_Batch1.xlsx")
    batch2_excel_path = os.path.join(data_folder, "QPN_Batch2.xlsx")
    batch3_excel_path = os.path.join(data_folder, "QPN_Batch3.xlsx")

    batch1_workbook = openpyxl.load_workbook(batch1_excel_path)
    batch1_sheet = batch1_workbook.active
    batch2_workbook = openpyxl.load_workbook(batch2_excel_path)
    batch2_sheet = batch2_workbook["Demographic"]
    batch3_workbook = openpyxl.load_workbook(batch3_excel_path)
    batch3_sheet = batch3_workbook.active

    all_patients = {}
    skip_dirs = {"noise", "rir", "test"}

    for dataset in os.listdir(data_folder):
        dataset_path = os.path.join(data_folder, dataset)

        if os.path.isfile(dataset_path) or dataset in skip_dirs:
            continue

        batch1_files = glob.glob(os.path.join(dataset_path, "Batch1", "*.wav"))
        batch2_files = glob.glob(os.path.join(dataset_path, "Batch2", "*.wav"))
        batch3_files = glob.glob(os.path.join(dataset_path, "Batch3", "*.wav"))

        all_patients.update(get_patient_traits(batch1_files, batch1_sheet, "Batch1"))
        all_patients.update(get_patient_traits(batch2_files, batch2_sheet, "Batch2"))
        all_patients.update(get_patient_traits(batch3_files, batch3_sheet, "Batch3"))

    return all_patients


def stratified_patient_split(path_type_dict):
    """
    Splits recordings into two halves by patient ID, stratified by ptype
    (Disease/Control), so each half has a similar class balance.

    Sorting by PID before splitting ensures the assignment is deterministic —
    the same patients will always land in the same split regardless of
    filesystem ordering.

    :param path_type_dict: dict of {path: patient_traits}
    :return: (split_a, split_b) — two dicts of {path: patient_traits}
    """
    # Group unique patient IDs by ptype
    ptype_to_pids = {}
    pid_to_paths = {}

    for path, traits in path_type_dict.items():
        pid = traits["pid"] if "pid" in traits else pathlib.Path(path).stem.split("_")[1]
        ptype = traits["ptype"]

        ptype_to_pids.setdefault(ptype, set()).add(pid)
        pid_to_paths.setdefault(pid, []).append(path)

    split_a_pids = set()
    split_b_pids = set()

    for ptype, pids in ptype_to_pids.items():
        sorted_pids = sorted(pids)           # Sort for determinism
        midpoint = len(sorted_pids) // 2
        split_a_pids.update(sorted_pids[:midpoint])
        split_b_pids.update(sorted_pids[midpoint:])

    def build_dict(pids):
        return {
            path: traits
            for path, traits in path_type_dict.items()
            for pid in [pathlib.Path(path).stem.split("_")[1]]
            if pid in pids
        }

    split_a = build_dict(split_a_pids)
    split_b = build_dict(split_b_pids)

    # Log the split composition for verification
    for name, pids in [("Split A", split_a_pids), ("Split B", split_b_pids)]:
        counts = {}
        for pid in pids:
            ptype = path_type_dict[pid_to_paths[pid][0]]["ptype"]
            counts[ptype] = counts.get(ptype, 0) + 1
        print(f"{name}: {len(pids)} patients — {counts}")

    return split_a, split_b


def get_patient_traits(files, sheet, batch):
    pids = [path.split("/")[-1].split("_")[1] for path in files]
    patients = {}

    for row in range(2, sheet.max_row + 1):
        pid = sheet.cell(row=row, column=1).value
        ptype = sheet.cell(row=row, column=2).value
        sex = sheet.cell(row=row, column=3).value
        l1 = sheet.cell(row=row, column=4).value
        age = sheet.cell(row=row, column=6).value

        if pid is not None and pid.rstrip() in pids:
            ptype = ptype.rstrip()
            sex = sex.rstrip()
            l1 = l1.rstrip()

            if ptype == "CTRL" or ptype == "control":
                ptype = "Control"
            elif ptype == "PD" or ptype == "patient":
                ptype = "Disease"
            else:
                print(f"Unknown key found: {ptype}")
                continue

            if l1 == "FR" or "French" in l1 or "Fench" in l1:
                l1 = "French"
            elif l1 == "EN" or "English" in l1:
                l1 = "English"
            else:
                l1 = "Other"

            patients[pid] = {"ptype": ptype, "sex": sex, "age": age, "l1": l1}

    updated_dict = {}
    for pid in patients:
        for path in files:
            if pid in path:
                updated_dict[path] = patients[pid]

    return updated_dict


def create_json(json_file, path_type_dict, chunk_size, overlap=None):
    hop_size = chunk_size / 2 if overlap is None else chunk_size - overlap
    json_dict = {}

    for audiofile in path_type_dict.keys():
        info_dict = path_type_dict[audiofile].copy()

        if "l1" in audiofile:
            continue

        audiopath = pathlib.Path(audiofile)
        uttid = audiopath.stem + "_" + audiopath.parent.name

        items = audiopath.stem.split("_")
        info_dict.update({"pid": items[1], "task": items[2], "lang": items[-1]})

        if info_dict["task"] not in ["a1", "a2", "a3", "a4", "vowel_repeat", "dpt", "recall", "repeat", "hbd", "read"]:
            info_dict["task"] = items[3]
        if info_dict["task"] in ["a1", "a2", "a3", "a4"]:
            info_dict["task"] = "vowel_repeat"
        if info_dict["lang"] not in ["en", "fr"]:
            info_dict["lang"] = "other"

        audioinfo = torchaudio.info(audiofile)
        duration = audioinfo.num_frames / audioinfo.sample_rate

        max_start = max(duration - hop_size, 1)
        for i, start in enumerate(np.arange(0, max_start, hop_size)):
            chunk_duration = min(chunk_size, duration - i * hop_size)
            json_dict[f"{uttid}_{i}"] = {
                "wav": audiofile,
                "start": start,
                "duration": chunk_duration,
                "info_dict": info_dict,
            }

    with open(json_file, mode="w") as json_f:
        json_dict = convert_to_python(json_dict)
        json.dump(json_dict, json_f, indent=2)
