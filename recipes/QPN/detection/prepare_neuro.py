import json
import os
import glob
import torchaudio
import openpyxl
import numpy as np
import pathlib


def convert_to_python(obj):
    """Recursively convert NumPy types to Python types for JSON serialization."""
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {k: convert_to_python(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_python(item) for item in obj]
    else:
        return obj


def prepare_neuro(
    data_folder, train_annotation, test_annotation, valid_annotation, chunk_size):
    assert os.path.exists(data_folder), "Data folder not found"

    path_type_dict = get_path_type_dicts(data_folder)

    create_json(train_annotation, path_type_dict["train"], chunk_size, overlap=0)
    create_json(test_annotation, path_type_dict["pd_test"], chunk_size)
    create_json(valid_annotation, path_type_dict["valid"], chunk_size)

def get_path_type_dicts(data_folder):
    """
    Function that extracts the patient_type and the path for each recording.
    Takes data_folder path, returns dicts with path:patient_type for each set.

    :param data_folder: string
    :return: dicts
    """

    datasets = os.listdir(data_folder)
    batch1_excel_path = os.path.join(data_folder, "QPN_Batch1.xlsx")
    batch2_excel_path = os.path.join(data_folder, "QPN_Batch2.xlsx")
    batch3_excel_path = os.path.join(data_folder, "QPN_Batch3.xlsx")
    path_type_dict = {}

    # Load the Excel files
    batch1_workbook = openpyxl.load_workbook(batch1_excel_path)
    batch1_sheet = batch1_workbook.active
    batch2_workbook = openpyxl.load_workbook(batch2_excel_path)
    batch2_sheet = batch2_workbook["Demographic"]
    batch3_workbook = openpyxl.load_workbook(batch3_excel_path)
    batch3_sheet = batch3_workbook.active

    for dataset in datasets:
        dataset_path = os.path.join(data_folder, dataset)

        # Skip files
        if os.path.isfile(dataset_path):
            continue
        if dataset == "noise" or dataset == "rir":
            continue

        batch1_data_path = os.path.join(dataset_path, "Batch1")
        batch2_data_path = os.path.join(dataset_path, "Batch2")
        batch3_data_path = os.path.join(dataset_path, "Batch3")

        batch1_files = glob.glob(batch1_data_path + "/*.wav")
        batch2_files = glob.glob(batch2_data_path + "/*.wav")
        batch3_files = glob.glob(batch3_data_path + "/*.wav")

        batch1_patients = get_patient_traits(batch1_files, batch1_sheet, "Batch1")
        batch2_patients = get_patient_traits(batch2_files, batch2_sheet, "Batch2")
        batch3_patients = get_patient_traits(batch3_files, batch3_sheet, "Batch3")

        path_type_dict[dataset] = batch1_patients | batch2_patients | batch3_patients

    path_type_dict["pd_test"] = path_type_dict["test_fr"] | path_type_dict["test_en"]
    del path_type_dict["test_fr"]
    del path_type_dict["test_en"]

    return path_type_dict


def get_patient_traits(files, sheet, batch):
    pids = [path.split("/")[-1].split("_")[1] for path in files]
    patients = {}

    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
        pid = sheet.cell(row=row, column=1).value
        if batch == "Batch1" or batch == "Batch3":
            ptype = sheet.cell(row=row, column=2).value
        else:
            ptype = sheet.cell(row=row, column=4).value

        if batch == "Batch1" or batch == "Batch3":
            sex = sheet.cell(row=row, column=3).value
        else:
            sex = sheet.cell(row=row, column=5).value

        if batch == "Batch1" or batch == "Batch3":
            l1 = sheet.cell(row=row, column=4).value
        else:
            l1 = sheet.cell(row=row, column=6).value

        if batch == "Batch1":
            age = sheet.cell(row=row, column=5).value
        elif batch == "Batch3":
            age = sheet.cell(row=row, column=6).value
        else:
            age = 0

        # Check if the patient ID is in the recordings, if it is add to dict
        if pid is not None and pid.rstrip() in pids:

            # rstrip() everything
            ptype = ptype.rstrip()
            sex = sex.rstrip()
            l1 = l1.rstrip()

            # Refactor patient type
            if ptype == "CTRL" or ptype == "control":
                ptype = "HC"
            elif ptype == "PD" or ptype == "patient":
                ptype = "PD"
            else:
                print(f"Unknown key found: {ptype}")
                continue

            # Refactor language
            if l1 == "FR" or "French" in l1 or "Fench" in l1:
                l1 = "French"
            elif l1 == "EN" or "English" in l1:
                l1 = "English"
            else:
                l1 = "Other"

            # Save to dict
            patient_traits = {
                "ptype": ptype,
                "sex": sex,
                "age": age,
                "l1": l1,
            }
            patients[pid] = patient_traits

    # Change pids to paths
    updated_dict = {}
    for pid in patients:
        for path in files:
            if pid in path:
                updated_dict[path] = patients[pid]

    return updated_dict


def create_json(json_file, path_type_dict, chunk_size, overlap=None):
    hop_size = chunk_size / 2 if overlap is None else chunk_size - overlap
    json_dict = {}

    for audiofile in path_type_dict.keys():
        # Get info dict
        info_dict = path_type_dict[audiofile].copy()

        # Remove 'l1' files as they are duplicates
        if "l1" in audiofile:
            continue

        # Get utterance info from the file name
        audiopath = pathlib.Path(audiofile)
        uttid = audiopath.stem + "_" + audiopath.parent.name

        # Second and third items are the PID and task, last is usually the language
        items = audiopath.stem.split("_")
        info_dict.update({"pid": items[1], "task": items[2], "lang": items[-1]})

        # Corrections
        if info_dict["task"] in ["a1", "a2", "a3", "a4"]:
            info_dict["task"] = "vowel_repeat"
        if info_dict["lang"] not in ["en", "fr"]:
            info_dict["lang"] = "other"

        # Get duration
        audioinfo = torchaudio.info(audiofile)
        duration = audioinfo.num_frames / audioinfo.sample_rate

        max_start = max(duration - hop_size, 1)
        for i, start in enumerate(np.arange(0, max_start, hop_size)):
            chunk_duration = min(chunk_size, duration - i * hop_size)
            json_dict[f"{uttid}_{i}"] = {
                "wav": audiofile,
                "start": start,
                "duration": chunk_duration,
                "info_dict": info_dict,
            }

    # Writing the dictionary to the json file
    with open(json_file, mode="w") as json_f:
        json_dict = convert_to_python(json_dict)
        json.dump(json_dict, json_f, indent=2)
